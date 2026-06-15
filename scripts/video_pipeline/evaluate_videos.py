"""
evaluate_videos.py — Stage 4: VLM evaluation on extracted frames.

Sends every extracted frame to CLIP, BLIP-2, LLaVA, and Qwen via the
VLM Disaster Analyzer backend API. Collects per-frame predictions,
aggregates to per-video majority-vote predictions, and exports:

  evaluation/
    clip_results.csv             — per-video CLIP predictions
    blip2_results.csv            — per-video BLIP-2 predictions
    llava_results.csv            — per-video LLaVA predictions
    qwen_results.csv             — per-video Qwen predictions
    clip_frame_results.csv       — per-frame CLIP results (full detail)
    blip2_frame_results.csv      — per-frame BLIP-2 results
    llava_frame_results.csv      — per-frame LLaVA results
    qwen_frame_results.csv       — per-frame Qwen results
    Video_VLM_Comparison.xlsx    — all models side-by-side

Prerequisites
-------------
1. The VLM backend must be running and accessible:
      python backend/main.py
   OR set VLM_BACKEND_URL (or the VLM_BACKEND env var in config.py)
   to your ngrok tunnel:
      os.environ["VLM_BACKEND"] = "https://xxxx.ngrok-free.app"

2. For BLIP-2 and LLaVA evaluation, start the backend with:
      ACTIVE_MODELS=clip,blip2,llava,qwen python backend/main.py

3. Extract frames first:
      python extract_frames.py

Usage
-----
    python evaluate_videos.py                              # all 4 models
    python evaluate_videos.py --models clip qwen           # CLIP + Qwen only
    python evaluate_videos.py --backend https://xyz.ngrok  # custom URL
"""

from __future__ import annotations

import io
import sys
import logging
import argparse
from pathlib import Path
from collections import Counter
from typing import Optional

import requests
import pandas as pd
from PIL import Image
from tqdm import tqdm

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    FRAMES_ROOT, METADATA_DIR, EVALUATION_DIR,
    TARGET_CATEGORIES, VLM_BACKEND_URL, VLM_TIMEOUT_S,
    HEADER_BG, HEADER_FG, CORRECT_BG, WRONG_BG,
)

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

ALL_MODELS = ["clip", "blip2", "llava", "qwen"]

# Maps research category to the broad label CLIP typically returns.
# Used for lightweight "correct" check without exact string matching.
_CATEGORY_ALIASES: dict[str, set[str]] = {
    "flood":      {"water disaster", "flood", "flooded"},
    "wildfire":   {"wild fire", "wildfire", "fire", "on fire", "urban fire"},
    "earthquake": {"earthquake", "infrastructure damage"},
    "landslide":  {"landslide", "mudslide"},
    "cyclone":    {"cyclone", "water disaster", "tropical cyclone", "tornado"},
}


# ── API helpers ────────────────────────────────────────────────────────────────

def _to_jpeg_bytes(path: Path) -> bytes:
    """Load a frame image file and return JPEG-encoded bytes."""
    img = Image.open(path).convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return buf.getvalue()


def _call_model(frame_path: Path, model: str, backend_url: str) -> dict:
    """
    POST one frame to POST /{model} on the backend.
    Returns the parsed JSON response, or an error dict on failure.
    """
    url  = f"{backend_url.rstrip('/')}/predict/{model}"
    jpeg = _to_jpeg_bytes(frame_path)
    try:
        resp = requests.post(
            url,
            files={"file": ("frame.jpg", jpeg, "image/jpeg")},
            timeout=VLM_TIMEOUT_S,
        )
        resp.raise_for_status()
        return resp.json()
    except requests.Timeout:
        return {"error": f"Timeout ({VLM_TIMEOUT_S}s)"}
    except requests.ConnectionError:
        return {"error": f"Cannot connect to {backend_url} — is the backend running?"}
    except Exception as e:
        return {"error": str(e)}


# ── Per-model evaluation ───────────────────────────────────────────────────────

def evaluate_model(model: str, backend_url: str) -> pd.DataFrame:
    """
    Evaluate all extracted frames with one VLM model.
    Returns a per-frame DataFrame.
    """
    frame_csv = METADATA_DIR / "frame_manifest.csv"
    if not frame_csv.exists():
        raise FileNotFoundError(
            f"frame_manifest.csv not found. Run extract_frames.py first.\n"
            f"Expected: {frame_csv}"
        )

    frames_df = pd.read_csv(frame_csv)
    frames_df = frames_df[frames_df.get("status", pd.Series(["extracted"] * len(frames_df)))
                          .isin(["extracted", "cached"])].copy()

    records = []
    for _, row in tqdm(frames_df.iterrows(), total=len(frames_df),
                       desc=f"[{model.upper():5}]", unit="frame"):

        frame_path = Path(str(row.get("frame_path", "")))
        if not frame_path.exists():
            records.append({
                "video_stem":    row.get("video_stem"),
                "category":      row.get("category"),
                "frame_path":    str(frame_path),
                "frame_number":  row.get("frame_number"),
                "model":         model,
                "prediction":    "MISSING",
                "confidence":    None,
                "raw_response":  "",
                "error":         "frame file not found",
            })
            continue

        result  = _call_model(frame_path, model, backend_url)
        metrics = result.get("metrics", {})

        prediction = (
            metrics.get("disaster_type")
            or metrics.get("scene_description", "")[:80]
            or result.get("error", "N/A")
        )

        records.append({
            "video_stem":    row.get("video_stem"),
            "category":      row.get("category"),
            "frame_path":    str(frame_path),
            "frame_number":  row.get("frame_number"),
            "model":         model,
            "prediction":    prediction,
            "confidence":    metrics.get("confidence_score"),
            "raw_response":  str(
                metrics.get("raw_analysis")
                or metrics.get("scene_description", "")
            )[:200],
            "error":         result.get("error", ""),
        })

    return pd.DataFrame(records)


# ── Aggregation ────────────────────────────────────────────────────────────────

def _is_correct(prediction: str, category: str) -> bool:
    """Lightweight correctness check using alias matching."""
    pred_lower = prediction.lower()
    aliases    = _CATEGORY_ALIASES.get(category, {category})
    return any(alias in pred_lower for alias in aliases)


def aggregate_per_video(frame_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate per-frame predictions to per-video results using majority vote.
    Adds 'correct' column based on alias matching.
    """
    records = []
    for (video_stem, category), grp in frame_df.groupby(["video_stem", "category"]):
        preds = [p for p in grp["prediction"].tolist()
                 if p and p not in ("MISSING", "N/A", "")]
        majority   = Counter(preds).most_common(1)[0][0] if preds else "N/A"
        avg_conf   = grp["confidence"].mean()
        model_name = grp["model"].iloc[0] if not grp.empty else "unknown"

        records.append({
            "video_stem":          video_stem,
            "category":            category,
            "model":               model_name,
            "frames_evaluated":    len(grp),
            "majority_prediction": majority,
            "avg_confidence":      round(float(avg_conf), 2) if pd.notna(avg_conf) else None,
            "correct":             _is_correct(majority, category),
        })

    return pd.DataFrame(records)


# ── Excel comparison workbook ──────────────────────────────────────────────────

def _export_comparison_excel(model_video_dfs: dict[str, pd.DataFrame]) -> Path:
    """
    Write Video_VLM_Comparison.xlsx:
      - One sheet per model (CLIP, BLIP2, LLAVA, QWEN)
      - Comparison sheet: video × model pivot
      - Summary sheet: accuracy + confidence per model
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    out = EVALUATION_DIR / "Video_VLM_Comparison.xlsx"
    EVALUATION_DIR.mkdir(parents=True, exist_ok=True)

    fill_h   = PatternFill("solid", fgColor=HEADER_BG)
    font_h   = Font(color=HEADER_FG, bold=True)
    fill_ok  = PatternFill("solid", fgColor=CORRECT_BG)
    fill_bad = PatternFill("solid", fgColor=WRONG_BG)
    align_c  = Alignment(horizontal="center")

    def _style(ws, df: pd.DataFrame) -> None:
        for cell in ws[1]:
            cell.fill = fill_h; cell.font = font_h; cell.alignment = align_c
        for ci, col in enumerate(df.columns, start=1):
            max_len = max(len(str(col)),
                          df[col].astype(str).str.len().max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Per-model sheets
        for model, df in model_video_dfs.items():
            sname = model.upper()
            df.to_excel(writer, sheet_name=sname, index=False)
            ws = writer.sheets[sname]
            _style(ws, df)
            if "correct" in df.columns:
                c_col = df.columns.get_loc("correct") + 1
                for ri, correct in enumerate(df["correct"], start=2):
                    ws.cell(ri, c_col).fill = fill_ok if correct else fill_bad

        # Comparison pivot: video_stem × model → majority_prediction
        all_video = pd.concat([
            df[["video_stem", "category", "majority_prediction", "correct"]]
            .assign(model=m)
            for m, df in model_video_dfs.items()
        ], ignore_index=True)

        pivot = all_video.pivot_table(
            index=["video_stem", "category"],
            columns="model",
            values="majority_prediction",
            aggfunc="first",
        ).reset_index()
        pivot.to_excel(writer, sheet_name="Comparison", index=False)
        _style(writer.sheets["Comparison"], pivot)

        # Summary: accuracy + avg confidence per model
        summary_rows = []
        for model, df in model_video_dfs.items():
            acc      = df["correct"].mean() * 100 if not df.empty else 0
            avg_conf = df["avg_confidence"].mean() if "avg_confidence" in df.columns else None
            summary_rows.append({
                "Model":            model.upper(),
                "Videos Evaluated": len(df),
                "Correct":          int(df["correct"].sum()) if not df.empty else 0,
                "Accuracy (%)":     round(acc, 1),
                "Avg Confidence":   round(float(avg_conf), 2)
                                    if avg_conf is not None and pd.notna(avg_conf) else "N/A",
            })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        _style(writer.sheets["Summary"], summary_df)

    log.info(f"Comparison workbook → {out}")
    return out


# ── Main pipeline ──────────────────────────────────────────────────────────────

def run_evaluation(
    models:      list[str] = ALL_MODELS,
    backend_url: str       = VLM_BACKEND_URL,
) -> None:
    """
    Evaluate all 75 videos with the specified VLM models.

    Args:
        models      : Models to evaluate. Subset of ["clip", "blip2", "llava", "qwen"].
        backend_url : VLM Disaster Analyzer backend URL.
    """
    EVALUATION_DIR.mkdir(parents=True, exist_ok=True)

    log.info("=" * 60)
    log.info(f"  VLM Evaluation | Models: {models}")
    log.info(f"  Backend : {backend_url}")
    log.info("=" * 60)

    model_video_dfs: dict[str, pd.DataFrame] = {}

    for model in models:
        log.info(f"\n── {model.upper()} ─────────────────────────────────────────────")
        try:
            frame_df = evaluate_model(model, backend_url)
            video_df = aggregate_per_video(frame_df)

            # Save CSVs
            frame_df.to_csv(EVALUATION_DIR / f"{model}_frame_results.csv", index=False)
            video_df.to_csv(EVALUATION_DIR / f"{model}_results.csv",       index=False)
            model_video_dfs[model] = video_df

            n_correct = video_df["correct"].sum() if not video_df.empty else 0
            n_total   = len(video_df)
            acc       = n_correct / n_total * 100 if n_total else 0
            log.info(f"  Accuracy: {acc:.1f}%  ({n_correct}/{n_total})")

        except FileNotFoundError as e:
            log.error(f"  {e}")
        except Exception as e:
            log.error(f"  Error evaluating {model}: {e}", exc_info=True)

    if model_video_dfs:
        _export_comparison_excel(model_video_dfs)
        log.info("\n── Evaluation complete ────────────────────────────────────")
        log.info(f"  CSVs   → {EVALUATION_DIR}")
        log.info(f"  Excel  → {EVALUATION_DIR / 'Video_VLM_Comparison.xlsx'}")
    else:
        log.warning("No models evaluated successfully.")


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    p = argparse.ArgumentParser(
        description="Evaluate 75 research videos with VLM models",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--models",  nargs="+", default=ALL_MODELS, choices=ALL_MODELS,
                   help="Models to use for evaluation")
    p.add_argument("--backend", default=VLM_BACKEND_URL,
                   help="VLM Disaster Analyzer backend URL")
    args = p.parse_args()
    run_evaluation(models=args.models, backend_url=args.backend)
