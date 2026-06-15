"""
generate_excel.py — Stage 3: manifest CSVs and statistics Excel workbook.

Scans raw_videos/ and extracted_frames/ to produce:
  metadata/
    video_manifest.csv       — one row per video (75 rows)
    frame_manifest.csv       — one row per frame (75 × 8 = 600 rows)
    dataset_statistics.xlsx  — 4-sheet Excel workbook

Excel sheets
------------
  Video Manifest    — all 75 videos with OpenCV metadata
  Frame Manifest    — all extracted frames
  Category Summary  — per-category aggregates (video count, duration, size, FPS)
  Statistics        — dataset-level key-value summary

Usage
-----
    python generate_excel.py
    python generate_excel.py --skip-frames   # omit Frame Manifest sheet
"""

from __future__ import annotations

import os
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime, timedelta

import cv2
import pandas as pd
from tqdm import tqdm

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    RAW_VIDEOS_ROOT, FRAMES_ROOT, METADATA_DIR,
    TARGET_CATEGORIES, HEADER_BG, HEADER_FG, ALT_ROW_BG,
)

log = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv", ".webm"}


# ── Video metadata extraction ──────────────────────────────────────────────────

def _video_metadata(video_path: Path) -> dict:
    """
    Open a video with OpenCV and extract stream-level metadata.
    Returns a dict with None values for fields that cannot be read.
    """
    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        return dict(frame_count=None, fps=None, duration_seconds=None,
                    width=None, height=None, resolution=None, corrupted=True)

    frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps    = cap.get(cv2.CAP_PROP_FPS)
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    cap.release()

    if fps <= 0 or frames <= 0:
        return dict(frame_count=frames, fps=fps, duration_seconds=None,
                    width=width, height=height, resolution=None, corrupted=True)

    return dict(
        frame_count      = frames,
        fps              = round(fps, 2),
        duration_seconds = round(frames / fps, 2),
        width            = width,
        height           = height,
        resolution       = f"{width}x{height}",
        corrupted        = False,
    )


# ── Build manifests ────────────────────────────────────────────────────────────

def build_video_manifest() -> pd.DataFrame:
    """
    Scan raw_videos/{category}/ and build the video manifest.

    Each row contains:
      video_id, category, file_name, youtube_id, clip_start_s, clip_end_s,
      file_size_mb, frame_count, fps, duration_seconds, width, height,
      resolution, corrupted

    Saves to metadata/video_manifest.csv.
    """
    METADATA_DIR.mkdir(parents=True, exist_ok=True)
    rows   = []
    vid_id = 1

    for category in TARGET_CATEGORIES:
        vid_dir = RAW_VIDEOS_ROOT / category
        if not vid_dir.exists():
            log.warning(f"  [{category}] Folder not found: {vid_dir}")
            continue

        videos = sorted(p for p in vid_dir.iterdir()
                        if p.suffix.lower() in VIDEO_EXTENSIONS)

        for vpath in tqdm(videos, desc=f"Manifest {category:12}", unit="video", leave=False):
            meta   = _video_metadata(vpath)
            parts  = vpath.stem.split("_")

            # File name format: {yt_id}_{start}_{end}.mp4
            yt_id  = parts[0]                       if parts          else ""
            t_s    = parts[1]                       if len(parts) > 1 else ""
            t_e    = parts[2]                       if len(parts) > 2 else ""

            rows.append({
                "video_id":         f"VID-{vid_id:04d}",
                "category":         category,
                "file_name":        vpath.name,
                "video_path":       str(vpath),
                "youtube_id":       yt_id,
                "clip_start_s":     t_s,
                "clip_end_s":       t_e,
                "file_size_mb":     round(vpath.stat().st_size / (1024 ** 2), 3),
                **meta,
            })
            vid_id += 1

    manifest = pd.DataFrame(rows)
    out = METADATA_DIR / "video_manifest.csv"
    manifest.to_csv(out, index=False)
    log.info(f"Video manifest: {len(manifest)} videos → {out}")
    return manifest


def build_frame_manifest() -> pd.DataFrame:
    """
    Re-scan extracted_frames/ and rebuild frame_manifest.csv.
    Also returns the existing manifest from metadata/ if available and no
    new frames are found (avoids unnecessary disk traversal).
    """
    existing_csv = METADATA_DIR / "frame_manifest.csv"
    frame_files  = list(FRAMES_ROOT.rglob(f"*.jpg")) + list(FRAMES_ROOT.rglob("*.png"))

    if not frame_files and existing_csv.exists():
        log.info(f"Frame manifest loaded from cache: {existing_csv}")
        return pd.read_csv(existing_csv)

    rows = []
    for frame_path in sorted(frame_files):
        parts = frame_path.parts
        # Expected structure: .../extracted_frames/{category}/{video_stem}/frame_001.jpg
        try:
            category   = frame_path.parent.parent.name
            video_stem = frame_path.parent.name
        except Exception:
            category = video_stem = "unknown"

        rows.append({
            "video_stem":    video_stem,
            "category":      category,
            "frame_path":    str(frame_path),
            "frame_filename": frame_path.name,
            "frame_number":  int(frame_path.stem.split("_")[-1])
                             if frame_path.stem[-1].isdigit() else 0,
        })

    manifest = pd.DataFrame(rows)
    manifest.to_csv(existing_csv, index=False)
    log.info(f"Frame manifest: {len(manifest)} frames → {existing_csv}")
    return manifest


# ── Aggregation helpers ────────────────────────────────────────────────────────

def _fmt_duration(seconds: float) -> str:
    if not seconds or pd.isna(seconds):
        return "N/A"
    return str(timedelta(seconds=int(seconds)))


def build_category_summary(manifest: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-category statistics from the video manifest."""
    if manifest.empty:
        return pd.DataFrame()

    healthy = manifest[~manifest.get("corrupted", pd.Series([False] * len(manifest)))]

    agg = (
        healthy.groupby("category")
        .agg(
            total_videos      = ("video_id",         "count"),
            total_duration_s  = ("duration_seconds", "sum"),
            avg_duration_s    = ("duration_seconds", "mean"),
            min_duration_s    = ("duration_seconds", "min"),
            max_duration_s    = ("duration_seconds", "max"),
            total_frames      = ("frame_count",      "sum"),
            avg_fps           = ("fps",               "mean"),
            total_size_mb     = ("file_size_mb",      "sum"),
        )
        .reindex(TARGET_CATEGORIES)
        .fillna(0)
        .reset_index()
    )

    agg["total_duration"]  = agg["total_duration_s"].apply(_fmt_duration)
    agg["avg_duration_s"]  = agg["avg_duration_s"].round(1)
    agg["avg_fps"]         = agg["avg_fps"].round(2)
    agg["total_size_mb"]   = agg["total_size_mb"].round(2)
    agg["total_frames"]    = agg["total_frames"].astype(int)
    agg["total_videos"]    = agg["total_videos"].astype(int)

    return agg[[
        "category", "total_videos", "total_duration", "avg_duration_s",
        "min_duration_s", "max_duration_s", "total_frames", "avg_fps", "total_size_mb",
    ]].rename(columns={
        "category":       "Category",
        "total_videos":   "Total Videos",
        "total_duration": "Total Duration",
        "avg_duration_s": "Avg Duration (s)",
        "min_duration_s": "Min Duration (s)",
        "max_duration_s": "Max Duration (s)",
        "total_frames":   "Total Frames",
        "avg_fps":        "Avg FPS",
        "total_size_mb":  "Total Size (MB)",
    })


def build_statistics(manifest: pd.DataFrame) -> pd.DataFrame:
    """Build a key-value dataset statistics table."""
    healthy     = manifest[~manifest.get("corrupted", pd.Series([False] * len(manifest)))]
    total_dur   = healthy["duration_seconds"].sum()  if not healthy.empty else 0
    total_size  = manifest["file_size_mb"].sum()
    avg_fps     = healthy["fps"].mean()              if not healthy.empty else 0
    top_res     = (healthy["resolution"].value_counts().idxmax()
                   if not healthy.empty and "resolution" in healthy.columns else "N/A")

    rows = [
        ("Dataset Name",              "VIDI Research Subset — ISRO Disaster Analysis"),
        ("Source",                    "https://github.com/vididataset/VIDI"),
        ("Total Videos",              len(manifest)),
        ("Healthy Videos",            len(healthy)),
        ("Corrupted Videos",          len(manifest) - len(healthy)),
        ("Total Categories",          manifest["category"].nunique() if not manifest.empty else 0),
        ("Categories",                ", ".join(TARGET_CATEGORIES)),
        ("Videos per Category",       len(manifest) // max(manifest["category"].nunique(), 1)
                                      if not manifest.empty else 0),
        ("Total Duration",            _fmt_duration(total_dur)),
        ("Total Duration (seconds)",  round(total_dur, 1)),
        ("Avg Video Duration (s)",    round(healthy["duration_seconds"].mean(), 1)
                                      if not healthy.empty else 0),
        ("Total Frames",              int(healthy["frame_count"].sum()) if not healthy.empty else 0),
        ("Avg FPS",                   round(avg_fps, 2)),
        ("Most Common Resolution",    top_res),
        ("Total Dataset Size (MB)",   round(total_size, 2)),
        ("Total Dataset Size (GB)",   round(total_size / 1024, 3)),
        ("Generated",                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


# ── Excel export ───────────────────────────────────────────────────────────────

def export_excel(
    video_manifest:    pd.DataFrame,
    category_summary:  pd.DataFrame,
    statistics:        pd.DataFrame,
    frame_manifest:    Optional[pd.DataFrame] = None,
) -> Path:
    """
    Write all sheets to metadata/dataset_statistics.xlsx.
    Returns the output path.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    out = METADATA_DIR / "dataset_statistics.xlsx"

    fill_h  = PatternFill("solid", fgColor=HEADER_BG)
    font_h  = Font(color=HEADER_FG, bold=True)
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _style(ws, df: pd.DataFrame) -> None:
        """Apply header styling and auto-fit column widths."""
        for cell in ws[1]:
            cell.fill      = fill_h
            cell.font      = font_h
            cell.alignment = align_c
        for col_i, col_name in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col_name)),
                df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
            )
            ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 4, 55)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        video_manifest.to_excel(   writer, sheet_name="Video Manifest",   index=False)
        category_summary.to_excel( writer, sheet_name="Category Summary", index=False)
        statistics.to_excel(       writer, sheet_name="Statistics",       index=False)

        if frame_manifest is not None and not frame_manifest.empty:
            frame_manifest.to_excel(writer, sheet_name="Frame Manifest",  index=False)

        _style(writer.sheets["Video Manifest"],   video_manifest)
        _style(writer.sheets["Category Summary"], category_summary)
        _style(writer.sheets["Statistics"],       statistics)
        if frame_manifest is not None and not frame_manifest.empty:
            _style(writer.sheets["Frame Manifest"], frame_manifest)

    log.info(f"Excel workbook → {out}")
    return out


# ── Main ───────────────────────────────────────────────────────────────────────

def run_generate(skip_frames: bool = False) -> None:
    """Run the full manifest + statistics generation pipeline."""
    from typing import Optional
    METADATA_DIR.mkdir(parents=True, exist_ok=True)

    log.info("Building video manifest...")
    vm = build_video_manifest()

    log.info("Building frame manifest...")
    fm = None if skip_frames else build_frame_manifest()

    log.info("Building category summary...")
    cs = build_category_summary(vm)

    log.info("Building dataset statistics...")
    st = build_statistics(vm)

    log.info("Exporting Excel workbook...")
    export_excel(vm, cs, st, fm)

    log.info("\n── Output summary ────────────────────────────────────────")
    log.info(f"  video_manifest.csv      → {METADATA_DIR / 'video_manifest.csv'}")
    if not skip_frames:
        log.info(f"  frame_manifest.csv      → {METADATA_DIR / 'frame_manifest.csv'}")
    log.info(f"  dataset_statistics.xlsx → {METADATA_DIR / 'dataset_statistics.xlsx'}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(
        description="Generate video manifest and dataset statistics Excel",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--skip-frames", action="store_true",
                   help="Omit Frame Manifest sheet from Excel (faster for large datasets)")
    args = p.parse_args()
    run_generate(skip_frames=args.skip_frames)


# Make Optional importable from the module level for export_excel
from typing import Optional
