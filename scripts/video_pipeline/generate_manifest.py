# =============================================================================
# VLM Disaster Analyzer — Video Dataset Manifest Generator
# =============================================================================
# Run this script in Google Colab to scan your real disaster video dataset
# and produce a multi-sheet Excel manifest (video_dataset_manifest.xlsx).
#
# Prerequisites:
#   1. Mount Google Drive (cell 1 below).
#   2. Set DATASET_ROOT to the folder containing your class sub-folders.
#   3. Run all cells top to bottom.
#
# No synthetic data is generated. Every row in the output comes from an
# actual video file found on disk.
# =============================================================================


# ── Cell 1: Install dependencies ──────────────────────────────────────────────
# (openpyxl is the only package not pre-installed on Colab)

"""
!pip install -q openpyxl
"""


# ── Cell 2: Mount Google Drive ────────────────────────────────────────────────

"""
from google.colab import drive
drive.mount('/content/drive', force_remount=False)
print("Drive mounted at /content/drive")
"""


# ── Cell 3: Imports ───────────────────────────────────────────────────────────

import os
import sys
import time
import traceback
from pathlib import Path
from datetime import timedelta

import cv2
import numpy as np
import pandas as pd
from tqdm import tqdm

print("Imports OK")
print(f"OpenCV  : {cv2.__version__}")
print(f"Pandas  : {pd.__version__}")


# ── Cell 4: Configuration ─────────────────────────────────────────────────────

# ---------------------------------------------------------------------------
# Edit DATASET_ROOT to point to your DisasterVideoDataset folder.
#
# Examples:
#   Google Drive (after mounting):
#       DATASET_ROOT = "/content/drive/MyDrive/DisasterVideoDataset"
#
#   Colab local (uploaded via Files panel):
#       DATASET_ROOT = "/content/DisasterVideoDataset"
#
#   Local machine (running outside Colab):
#       DATASET_ROOT = r"C:/Users/ujjes/OneDrive/Desktop/DisasterVideoDataset"
# ---------------------------------------------------------------------------

DATASET_ROOT = "/content/drive/MyDrive/DisasterVideoDataset"

# Where to save the Excel file.
# Defaults to the dataset root so it stays next to the data on Drive.
OUTPUT_PATH = os.path.join(DATASET_ROOT, "video_dataset_manifest.xlsx")

# Recognised video extensions (lower-case).
VIDEO_EXTENSIONS = {".mp4", ".avi", ".mov", ".mkv"}

# ---------------------------------------------------------------------------
# Validate the dataset root exists before doing any work.
# ---------------------------------------------------------------------------
if not os.path.isdir(DATASET_ROOT):
    print(f"\n[ERROR] Dataset root not found:\n  {DATASET_ROOT}")
    print("\nFix: set DATASET_ROOT to the correct path and re-run this cell.")
    sys.exit(1)

print(f"Dataset root : {DATASET_ROOT}")
print(f"Output path  : {OUTPUT_PATH}")


# ── Cell 5: Discovery — scan all video files ──────────────────────────────────

def discover_videos(root: str, extensions: set[str]) -> list[dict]:
    """
    Recursively walk *root* and return one dict per video file found.

    The immediate sub-folder name is used as the disaster-type label.
    Nested sub-folders (e.g. root/Flood/clips/video.mp4) are included;
    their label is still the first-level folder name ("Flood").

    Returns
    -------
    list of {"video_path": str, "disaster_type": str, "file_name": str}
    """
    records = []

    # Sort so the order is deterministic across runs.
    for entry in sorted(os.scandir(root), key=lambda e: e.name):
        if not entry.is_dir():
            continue

        disaster_type = entry.name  # e.g. "Earthquake"

        for dirpath, _dirs, files in os.walk(entry.path):
            for fname in sorted(files):
                if os.path.splitext(fname)[1].lower() in extensions:
                    records.append({
                        "video_path":   os.path.join(dirpath, fname),
                        "disaster_type": disaster_type,
                        "file_name":    fname,
                    })

    return records


print("Scanning dataset...")
video_records = discover_videos(DATASET_ROOT, VIDEO_EXTENSIONS)

# Count per class for a quick preview.
class_counts: dict[str, int] = {}
for r in video_records:
    class_counts[r["disaster_type"]] = class_counts.get(r["disaster_type"], 0) + 1

print(f"\nFound {len(video_records)} video file(s) across {len(class_counts)} class(es):\n")
for cls, cnt in sorted(class_counts.items()):
    print(f"  {cls:<20} {cnt} video(s)")

if not video_records:
    print("\n[WARN] No videos found. Check DATASET_ROOT and make sure the class "
          "sub-folders contain .mp4 / .avi / .mov / .mkv files.")


# ── Cell 6: Metadata extraction ───────────────────────────────────────────────

def extract_metadata(video_path: str) -> dict:
    """
    Open a video file with OpenCV and extract per-file metadata.

    Returns a dict with all numeric fields.
    Returns None values (and sets "corrupted" = True) when the file
    cannot be opened or has zero frames — the caller logs these separately.
    """
    result = {
        "duration_seconds": None,
        "fps":              None,
        "frame_count":      None,
        "width":            None,
        "height":           None,
        "resolution":       None,
        "file_size_mb":     None,
        "corrupted":        False,
        "error_message":    "",
    }

    # File size is available even if the video codec is broken.
    try:
        result["file_size_mb"] = round(os.path.getsize(video_path) / (1024 ** 2), 3)
    except OSError as exc:
        result["corrupted"]     = True
        result["error_message"] = f"os.path.getsize failed: {exc}"
        return result

    cap = cv2.VideoCapture(video_path)

    if not cap.isOpened():
        result["corrupted"]     = True
        result["error_message"] = "cv2.VideoCapture could not open file"
        return result

    try:
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps         = cap.get(cv2.CAP_PROP_FPS)
        width       = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height      = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

        # Some containers report 0 FPS / 0 frames — treat as corrupted.
        if fps <= 0 or frame_count <= 0 or width <= 0 or height <= 0:
            result["corrupted"]     = True
            result["error_message"] = (
                f"Invalid stream properties: "
                f"fps={fps}, frames={frame_count}, w={width}, h={height}"
            )
            return result

        duration = round(frame_count / fps, 3)

        result.update({
            "frame_count":      frame_count,
            "fps":              round(fps, 3),
            "duration_seconds": duration,
            "width":            width,
            "height":           height,
            "resolution":       f"{width}x{height}",
        })

    except Exception as exc:
        result["corrupted"]     = True
        result["error_message"] = f"Metadata read error: {exc}"

    finally:
        cap.release()

    return result


# ── Cell 7: Build the manifest DataFrame ──────────────────────────────────────

def build_manifest(records: list[dict]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Iterate over discovered video records, extract metadata, and assemble two
    DataFrames:

      manifest_df  — one row per video (healthy + corrupted, clearly flagged)
      corrupt_df   — corrupted videos only, for separate review

    A monotonic Video ID (VID-0001, VID-0002, …) is assigned in discovery order.
    """
    rows   = []
    errors = []

    for idx, rec in enumerate(tqdm(records, desc="Extracting metadata", unit="video"), start=1):
        video_path   = rec["video_path"]
        disaster     = rec["disaster_type"]
        fname        = rec["file_name"]
        video_id     = f"VID-{idx:04d}"

        meta = extract_metadata(video_path)

        row = {
            "video_id":         video_id,
            "file_name":        fname,
            "video_path":       video_path,
            "disaster_type":    disaster,
            "duration_seconds": meta["duration_seconds"],
            "fps":              meta["fps"],
            "frame_count":      meta["frame_count"],
            "width":            meta["width"],
            "height":           meta["height"],
            "resolution":       meta["resolution"],
            "file_size_mb":     meta["file_size_mb"],
            "corrupted":        meta["corrupted"],
        }
        rows.append(row)

        if meta["corrupted"]:
            errors.append({
                "video_id":      video_id,
                "file_name":     fname,
                "video_path":    video_path,
                "disaster_type": disaster,
                "file_size_mb":  meta["file_size_mb"],
                "error_message": meta["error_message"],
            })

    manifest_df = pd.DataFrame(rows, columns=[
        "video_id", "file_name", "video_path", "disaster_type",
        "duration_seconds", "fps", "frame_count",
        "width", "height", "resolution",
        "file_size_mb", "corrupted",
    ])

    corrupt_df = pd.DataFrame(errors, columns=[
        "video_id", "file_name", "video_path", "disaster_type",
        "file_size_mb", "error_message",
    ])

    return manifest_df, corrupt_df


print("Building manifest — this reads every video file header...")
t_start = time.perf_counter()

manifest_df, corrupt_df = build_manifest(video_records)

t_elapsed = time.perf_counter() - t_start
healthy   = manifest_df[~manifest_df["corrupted"]]
print(f"\nDone in {t_elapsed:.1f}s")
print(f"  Total processed : {len(manifest_df)}")
print(f"  Healthy         : {len(healthy)}")
print(f"  Corrupted       : {len(corrupt_df)}")

manifest_df.head(5)


# ── Cell 8: Build Class Summary sheet ────────────────────────────────────────

def format_duration(total_seconds: float) -> str:
    """Convert seconds to HH:MM:SS string."""
    if pd.isna(total_seconds):
        return "N/A"
    td = timedelta(seconds=int(total_seconds))
    return str(td)


def build_class_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate per-disaster-type statistics from the healthy video subset.
    Only rows where corrupted == False are included in the numeric columns.
    """
    healthy = df[~df["corrupted"]].copy()

    # Videos per class (includes corrupted count for transparency).
    all_counts = (
        df.groupby("disaster_type")
        .size()
        .rename("total_videos")
    )
    corrupt_counts = (
        df[df["corrupted"]].groupby("disaster_type")
        .size()
        .rename("corrupted_videos")
    )

    agg = (
        healthy.groupby("disaster_type")
        .agg(
            healthy_videos    = ("video_id",         "count"),
            total_duration_s  = ("duration_seconds", "sum"),
            avg_duration_s    = ("duration_seconds", "mean"),
            min_duration_s    = ("duration_seconds", "min"),
            max_duration_s    = ("duration_seconds", "max"),
            total_frames      = ("frame_count",      "sum"),
            avg_fps           = ("fps",               "mean"),
            total_size_mb     = ("file_size_mb",      "sum"),
        )
    )

    summary = (
        agg
        .join(all_counts,     how="outer")
        .join(corrupt_counts, how="left")
        .fillna({"corrupted_videos": 0, "healthy_videos": 0})
    )

    # Human-readable duration columns.
    summary["total_duration"]   = summary["total_duration_s"].apply(format_duration)
    summary["avg_duration"]     = summary["avg_duration_s"].apply(
        lambda s: f"{s:.1f}s" if pd.notna(s) else "N/A"
    )
    summary["avg_fps"]          = summary["avg_fps"].round(2)
    summary["total_size_mb"]    = summary["total_size_mb"].round(2)
    summary["total_frames"]     = summary["total_frames"].fillna(0).astype(int)
    summary["total_videos"]     = summary["total_videos"].astype(int)
    summary["corrupted_videos"] = summary["corrupted_videos"].astype(int)
    summary["healthy_videos"]   = summary["healthy_videos"].astype(int)

    summary = summary.reset_index().rename(columns={"disaster_type": "Disaster Type"})

    return summary[[
        "Disaster Type", "total_videos", "healthy_videos", "corrupted_videos",
        "total_duration", "avg_duration",
        "min_duration_s", "max_duration_s",
        "total_frames", "avg_fps",
        "total_size_mb",
    ]].rename(columns={
        "total_videos":     "Total Videos",
        "healthy_videos":   "Healthy Videos",
        "corrupted_videos": "Corrupted Videos",
        "total_duration":   "Total Duration",
        "avg_duration":     "Avg Duration",
        "min_duration_s":   "Min Duration (s)",
        "max_duration_s":   "Max Duration (s)",
        "total_frames":     "Total Frames",
        "avg_fps":          "Avg FPS",
        "total_size_mb":    "Total Size (MB)",
    })


class_summary_df = build_class_summary(manifest_df)
print("Class Summary:")
print(class_summary_df.to_string(index=False))


# ── Cell 9: Build Dataset Statistics sheet ────────────────────────────────────

def build_dataset_statistics(manifest: pd.DataFrame, corrupt: pd.DataFrame) -> pd.DataFrame:
    """
    Produce a single-column key-value statistics table for the Statistics sheet.
    """
    healthy = manifest[~manifest["corrupted"]]

    total_duration_s  = healthy["duration_seconds"].sum()
    avg_duration_s    = healthy["duration_seconds"].mean()
    total_size_mb     = manifest["file_size_mb"].sum()
    resolutions       = healthy["resolution"].value_counts()
    top_res           = resolutions.idxmax() if not resolutions.empty else "N/A"
    fps_vals          = healthy["fps"].dropna()
    avg_fps           = fps_vals.mean() if not fps_vals.empty else float("nan")

    stats = [
        ("Total Videos",               len(manifest)),
        ("Healthy Videos",             len(healthy)),
        ("Corrupted Videos",           len(corrupt)),
        ("Total Classes",              manifest["disaster_type"].nunique()),
        ("Classes Detected",           ", ".join(sorted(manifest["disaster_type"].unique()))),
        ("Total Dataset Duration",     format_duration(total_duration_s)),
        ("Total Duration (seconds)",   round(total_duration_s, 2)),
        ("Average Video Length (s)",   round(avg_duration_s, 2) if pd.notna(avg_duration_s) else "N/A"),
        ("Total Frames (healthy)",     int(healthy["frame_count"].sum())),
        ("Average FPS",                round(avg_fps, 2) if pd.notna(avg_fps) else "N/A"),
        ("Most Common Resolution",     top_res),
        ("Total Dataset Size (MB)",    round(total_size_mb, 2)),
        ("Total Dataset Size (GB)",    round(total_size_mb / 1024, 3)),
        ("Manifest Generated",         pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Dataset Root",               DATASET_ROOT),
    ]

    return pd.DataFrame(stats, columns=["Metric", "Value"])


stats_df = build_dataset_statistics(manifest_df, corrupt_df)
print("\nDataset Statistics:")
print(stats_df.to_string(index=False))


# ── Cell 10: Export to Excel ──────────────────────────────────────────────────

def export_excel(
    output_path: str,
    manifest:    pd.DataFrame,
    class_summary: pd.DataFrame,
    statistics:  pd.DataFrame,
    corrupted:   pd.DataFrame,
) -> None:
    """
    Write four sheets to an Excel workbook:
      1. Manifest       — one row per video, all columns
      2. Class Summary  — aggregated per disaster type
      3. Statistics     — dataset-level key-value table
      4. Corrupted      — videos that could not be read (may be empty)

    Auto-fits column widths using openpyxl's column_dimensions.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    # Palette (matches VLM Disaster Analyzer warm ivory theme).
    HEADER_FILL   = PatternFill("solid", fgColor="543A14")
    HEADER_FONT   = Font(color="FFF0DC", bold=True)
    CORRUPT_FILL  = PatternFill("solid", fgColor="FF4444")
    CORRUPT_FONT  = Font(color="FFFFFF", bold=True)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: Manifest ──────────────────────────────────────────────
        manifest.to_excel(writer, sheet_name="Manifest", index=False)
        ws = writer.sheets["Manifest"]

        # Style header row.
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Highlight corrupted rows in red.
        corrupt_col = manifest.columns.get_loc("corrupted") + 1  # 1-indexed
        for row_idx, is_corrupt in enumerate(manifest["corrupted"], start=2):
            if is_corrupt:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor="FFD7D7")

        # Auto-fit column widths.
        for col_idx, col in enumerate(manifest.columns, start=1):
            max_len = max(
                len(str(col)),
                manifest[col].astype(str).str.len().max() if len(manifest) else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # ── Sheet 2: Class Summary ─────────────────────────────────────────
        class_summary.to_excel(writer, sheet_name="Class Summary", index=False)
        ws2 = writer.sheets["Class Summary"]
        for cell in ws2[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for col_idx, col in enumerate(class_summary.columns, start=1):
            max_len = max(
                len(str(col)),
                class_summary[col].astype(str).str.len().max() if len(class_summary) else 0,
            )
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

        # ── Sheet 3: Statistics ────────────────────────────────────────────
        statistics.to_excel(writer, sheet_name="Statistics", index=False)
        ws3 = writer.sheets["Statistics"]
        for cell in ws3[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 50

        # ── Sheet 4: Corrupted ─────────────────────────────────────────────
        if not corrupted.empty:
            corrupted.to_excel(writer, sheet_name="Corrupted", index=False)
            ws4 = writer.sheets["Corrupted"]
            for cell in ws4[1]:
                cell.fill = CORRUPT_FILL
                cell.font = CORRUPT_FONT
            for col_idx, col in enumerate(corrupted.columns, start=1):
                max_len = max(
                    len(str(col)),
                    corrupted[col].astype(str).str.len().max() if len(corrupted) else 0,
                )
                ws4.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)
        else:
            # Write an empty sheet so the tab always exists.
            pd.DataFrame(columns=corrupted.columns).to_excel(
                writer, sheet_name="Corrupted", index=False
            )

    print(f"\n[OK] Excel written to:\n  {output_path}")


export_excel(
    output_path   = OUTPUT_PATH,
    manifest      = manifest_df,
    class_summary = class_summary_df,
    statistics    = stats_df,
    corrupted     = corrupt_df,
)


# ── Cell 11: Final summary ────────────────────────────────────────────────────

healthy_count = len(manifest_df[~manifest_df["corrupted"]])
total_dur_s   = manifest_df.loc[~manifest_df["corrupted"], "duration_seconds"].sum()

print("\n" + "=" * 58)
print("  VLM DISASTER ANALYZER — MANIFEST GENERATION COMPLETE")
print("=" * 58)
print(f"  Total videos processed : {len(manifest_df)}")
print(f"  Healthy videos         : {healthy_count}")
print(f"  Corrupted / unreadable : {len(corrupt_df)}")
print(f"  Total classes found    : {manifest_df['disaster_type'].nunique()}")
print(f"  Classes                : {', '.join(sorted(manifest_df['disaster_type'].unique()))}")
print(f"  Total dataset duration : {format_duration(total_dur_s)}")
print(f"  Excel output path      : {OUTPUT_PATH}")
print("=" * 58)

if not corrupt_df.empty:
    print(f"\n[WARN] {len(corrupt_df)} corrupted file(s) logged in the 'Corrupted' sheet:")
    for _, row in corrupt_df.iterrows():
        print(f"  {row['video_id']}  {row['file_name']}")
        print(f"    Reason: {row['error_message']}")
