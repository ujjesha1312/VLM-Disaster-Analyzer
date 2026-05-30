#!/usr/bin/env python3
"""
batch_evaluate.py — Automated disaster image batch evaluation pipeline.

Traverses structured dataset folders, sends each image to the /predict/clip
endpoint, and exports prediction results + category summary to CSV and Excel.

Works with any folder depth — leaf directories (those containing images)
are discovered automatically, regardless of nesting level.

Usage (from project root, with server running on port 8000):

    python scripts/batch_evaluate.py

Custom paths:

    python scripts/batch_evaluate.py \\
        --dataset datasets/comprehensive_disaster_dataset \\
        --output  outputs \\
        --url     http://127.0.0.1:8000 \\
        --limit   50

Flags:
    --dataset PATH   Root folder to scan  (default: datasets/comprehensive_disaster_dataset)
    --output  PATH   Results folder       (default: outputs)
    --url     URL    Backend base URL     (default: http://127.0.0.1:8000)
    --limit   N      Max images per category (0 = no limit)
    --timeout N      Per-request timeout in seconds (default: 30)
"""

import argparse
import sys
import time
from collections import Counter
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency guard — clear error messages before traceback noise
# ---------------------------------------------------------------------------

try:
    import pandas as pd
except ImportError:
    sys.exit("[ERROR] pandas not installed.  Run: pip install pandas")

try:
    import requests
except ImportError:
    sys.exit("[ERROR] requests not installed.  Run: pip install requests")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] openpyxl not installed.  Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent

DEFAULT_DATASET_DIR = PROJECT_ROOT / "datasets" / "comprehensive_disaster_dataset"
DEFAULT_OUTPUT_DIR  = PROJECT_ROOT / "outputs"
DEFAULT_BASE_URL    = "http://127.0.0.1:8000"
CLIP_ENDPOINT       = "/predict/clip"

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff"}

EXTENSION_TO_MIME = {
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".bmp":  "image/bmp",
    ".webp": "image/webp",
    ".tiff": "image/tiff",
}

# Excel colour palette
_BLUE_DARK    = "1F4E79"
_BLUE_LIGHT   = "D6E4F0"
_GREEN_DARK   = "1B5E20"
_GREEN_LIGHT  = "E8F5E9"
_WHITE        = "FFFFFF"
_BORDER_GREY  = "B0BEC5"


# ---------------------------------------------------------------------------
# Dataset discovery — works with any folder depth
# ---------------------------------------------------------------------------

def _is_image(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS


def _format_label(folder_name: str) -> str:
    """Convert a folder name to a human-readable label.

    Examples:
        "Land_Slide"                   → "Land Slide"
        "Wild_Fire"                    → "Wild Fire"
        "Non_Damage_Buildings_Street"  → "Non Damage Buildings Street"
        "earthquake"                   → "Earthquake"
        "Earthquake"                   → "Earthquake"
    """
    return folder_name.replace("_", " ").strip().title()


def discover_leaf_categories(dataset_dir: Path, limit: int = 0) -> list[tuple[Path, str]]:
    """Recursively find all leaf directories (those that contain images).

    Returns a sorted list of (image_path, actual_label) pairs.
    The actual label is derived from the leaf folder name.

    If limit > 0, at most `limit` images are taken from each category.
    """
    if not dataset_dir.exists():
        sys.exit(
            f"\n[ERROR] Dataset not found: {dataset_dir}\n"
            f"  Create it with disaster subfolders, e.g.:\n"
            f"    {dataset_dir / 'flood'}/\n"
            f"    {dataset_dir / 'earthquake'}/\n"
            f"  Or point --dataset at your actual dataset root."
        )

    entries: list[tuple[Path, str]] = []

    for folder in sorted(dataset_dir.rglob("*")):
        if not folder.is_dir():
            continue
        images = sorted(f for f in folder.iterdir() if _is_image(f))
        if not images:
            continue  # not a leaf folder — no images directly inside

        label = _format_label(folder.name)
        if limit > 0:
            images = images[:limit]
        for img in images:
            entries.append((img, label))

    return entries


# ---------------------------------------------------------------------------
# API call
# ---------------------------------------------------------------------------

def call_clip(image_path: Path, endpoint_url: str, timeout: int = 30) -> dict:
    """POST an image to /predict/clip; return parsed JSON.

    Raises:
        requests.HTTPError       on 4xx / 5xx responses
        requests.ConnectionError when the backend is unreachable
        requests.Timeout         when the request exceeds `timeout` seconds
    """
    mime = EXTENSION_TO_MIME.get(image_path.suffix.lower(), "image/jpeg")
    with image_path.open("rb") as fh:
        resp = requests.post(
            endpoint_url,
            files={"file": (image_path.name, fh, mime)},
            timeout=timeout,
        )
    resp.raise_for_status()
    return resp.json()


# ---------------------------------------------------------------------------
# Inference loop
# ---------------------------------------------------------------------------

def run_evaluation(
    entries: list[tuple[Path, str]],
    endpoint_url: str,
    timeout: int,
) -> tuple[list[dict], list[dict]]:
    """Call the CLIP endpoint for every entry; return (results, errors)."""

    results: list[dict] = []
    errors:  list[dict] = []
    total = len(entries)
    idx_width = len(str(total))
    t_start = time.time()

    for idx, (image_path, actual_label) in enumerate(entries, start=1):
        # Build a short display path:  "Wild_Fire/img001.png"
        display = f"{image_path.parent.name}/{image_path.name}"

        # Estimate remaining time after first few requests
        elapsed = time.time() - t_start
        if idx > 3:
            avg_s    = elapsed / (idx - 1)
            remaining = avg_s * (total - idx + 1)
            eta_str  = f"  ETA {remaining/60:.0f}m{remaining%60:.0f}s"
        else:
            eta_str = ""

        print(
            f"  [{idx:>{idx_width}}/{total}]  {display:<50}",
            end="  ",
            flush=True,
        )

        try:
            data       = call_clip(image_path, endpoint_url, timeout)
            predicted  = data.get("prediction", "Unknown")
            confidence = float(data.get("confidence", 0.0))
            correct    = predicted.lower() == actual_label.lower()
            marker     = "✓" if correct else "✗"

            print(f"→ {predicted:<14}  {confidence:5.1f}%  {marker}{eta_str}")

            results.append({
                "Image Name":         image_path.name,
                "Category Path":      str(image_path.parent.relative_to(
                                          image_path.parent.parent.parent
                                      )),
                "Actual Disaster":    actual_label,
                "Predicted Disaster": predicted,
                "Confidence":         round(confidence, 2),
                "Correct":            correct,
            })

        except requests.exceptions.ConnectionError:
            print("  ✗  CONNECTION ERROR")
            errors.append({
                "Image Name": image_path.name,
                "Label":      actual_label,
                "Error":      "ConnectionError — is the server running?",
            })
        except requests.exceptions.Timeout:
            print(f"  ✗  TIMEOUT (>{timeout}s)")
            errors.append({
                "Image Name": image_path.name,
                "Label":      actual_label,
                "Error":      f"Timeout after {timeout}s",
            })
        except requests.exceptions.HTTPError as exc:
            print(f"  ✗  HTTP {exc.response.status_code}")
            errors.append({
                "Image Name": image_path.name,
                "Label":      actual_label,
                "Error":      str(exc),
            })
        except Exception as exc:
            print(f"  ✗  {exc}")
            errors.append({
                "Image Name": image_path.name,
                "Label":      actual_label,
                "Error":      str(exc),
            })

    return results, errors


# ---------------------------------------------------------------------------
# Summary computation
# ---------------------------------------------------------------------------

def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-category counts, average confidence, and accuracy."""
    grp = df.groupby("Actual Disaster", sort=True)
    summary = (
        grp.agg(
            **{
                "No. of Images":       ("Image Name", "count"),
                "Average Confidence":  ("Confidence", "mean"),
                "Correct Predictions": ("Correct",    "sum"),
            }
        )
        .reset_index()
        .rename(columns={"Actual Disaster": "Disaster Type"})
    )
    summary["Accuracy (%)"] = (
        summary["Correct Predictions"] / summary["No. of Images"] * 100
    ).round(1)
    summary["Average Confidence"] = summary["Average Confidence"].round(1)
    summary = summary.drop(columns=["Correct Predictions"])
    return summary[["Disaster Type", "No. of Images", "Average Confidence", "Accuracy (%)"]]


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

def _thin_border(color: str = _BORDER_GREY) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(bg_hex: str, font_hex: str = _WHITE) -> tuple:
    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color=font_hex, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _apply_formatting(ws, header_bg: str, even_bg: str, odd_bg: str = _WHITE) -> None:
    """Apply header style, alternating row colours, borders, and column widths."""
    h_fill, h_font, h_align = _header_style(header_bg)
    even_fill = PatternFill("solid", fgColor=even_bg)
    odd_fill  = PatternFill("solid", fgColor=odd_bg)
    border    = _thin_border()

    # Header row
    for cell in ws[1]:
        cell.fill      = h_fill
        cell.font      = h_font
        cell.alignment = h_align
        cell.border    = border
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for cell in row:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

    # Column widths
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = ws["A2"]


def write_results_excel(df: pd.DataFrame, path: Path) -> None:
    """Write results DataFrame to a formatted .xlsx file."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
    wb = openpyxl.load_workbook(path)
    _apply_formatting(wb["Results"], header_bg=_BLUE_DARK, even_bg=_BLUE_LIGHT)
    wb.save(path)


def write_summary_excel(df: pd.DataFrame, path: Path, overall: dict) -> None:
    """Write summary DataFrame + overall totals row to a formatted .xlsx file.

    overall keys expected: total_images, avg_confidence, accuracy, errors
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Summary", index=False)

    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    _apply_formatting(ws, header_bg=_GREEN_DARK, even_bg=_GREEN_LIGHT)

    # ── Blank separator row ───────────────────────────────────────────────────
    sep_row = ws.max_row + 2

    # ── OVERALL totals row ────────────────────────────────────────────────────
    totals_fill   = PatternFill("solid", fgColor="1B5E20")   # very dark green
    totals_font   = Font(bold=True, color=_WHITE, size=11)
    totals_align  = Alignment(horizontal="center", vertical="center")
    totals_border = _thin_border("145214")

    totals_values = [
        "OVERALL",
        overall["total_images"],
        round(overall["avg_confidence"], 1),
        round(overall["accuracy"], 1),
    ]

    for col_idx, value in enumerate(totals_values, start=1):
        cell = ws.cell(row=sep_row, column=col_idx, value=value)
        cell.fill      = totals_fill
        cell.font      = totals_font
        cell.alignment = totals_align
        cell.border    = totals_border
    ws.row_dimensions[sep_row].height = 22

    # ── Stats block two rows below totals ─────────────────────────────────────
    stats_row = sep_row + 2
    stats_font_label = Font(bold=True, color="1B5E20", size=10)
    stats_font_value = Font(color="333333", size=10)

    stats = [
        ("Total Images Processed", f"{overall['total_images']:,}"),
        ("Overall Accuracy",       f"{overall['accuracy']:.1f}%"),
        ("Overall Avg Confidence", f"{overall['avg_confidence']:.1f}%"),
        ("Total Errors",           str(overall["errors"])),
    ]

    for offset, (label, value) in enumerate(stats):
        ws.cell(row=stats_row + offset, column=1, value=label).font = stats_font_label
        ws.cell(row=stats_row + offset, column=2, value=value).font = stats_font_value

    wb.save(path)


# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Batch evaluate disaster images via the CLIP backend.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--dataset", type=Path, default=DEFAULT_DATASET_DIR,
        help="Root dataset folder to scan",
    )
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_OUTPUT_DIR,
        help="Folder to write results to",
    )
    parser.add_argument(
        "--url", default=DEFAULT_BASE_URL,
        help="Backend base URL",
    )
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Max images per category (0 = no limit)",
    )
    parser.add_argument(
        "--timeout", type=int, default=30,
        help="Per-request timeout in seconds",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

SEP = "─" * 62


def main() -> None:
    args = parse_args()

    endpoint_url = args.url.rstrip("/") + CLIP_ENDPOINT
    output_dir   = args.output

    # ── Header ───────────────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  VLM Disaster Analyzer — Batch Evaluation Pipeline")
    print(SEP)
    print(f"  Dataset  : {args.dataset}")
    print(f"  Output   : {output_dir}")
    print(f"  Endpoint : {endpoint_url}")
    if args.limit:
        print(f"  Limit    : {args.limit} images per category")
    print(SEP)

    # ── Discover images ───────────────────────────────────────────────────────
    print("\n  Scanning dataset...")
    entries = discover_leaf_categories(args.dataset, limit=args.limit)

    if not entries:
        sys.exit(
            f"\n[ERROR] No supported images found in {args.dataset}\n"
            f"  Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    category_counts = Counter(label for _, label in entries)
    print(f"\n  Discovered {len(entries):,} images across {len(category_counts)} categories:")
    for cat in sorted(category_counts):
        print(f"    {cat:<40} {category_counts[cat]:>6} images")

    # ── Server health check ───────────────────────────────────────────────────
    print(f"\n  Checking backend at {args.url} ...", end="  ", flush=True)
    try:
        r = requests.get(args.url.rstrip("/") + "/", timeout=5)
        r.raise_for_status()
        print("online ✓")
    except Exception:
        print("UNREACHABLE ✗")
        sys.exit(
            f"\n[ERROR] Backend is not running at {args.url}\n"
            f"  Start it with:  uvicorn backend.main:app --port 8000"
        )

    # Estimate total time
    est_seconds = len(entries) * 1.2   # rough 1.2s per image
    print(f"\n  Estimated duration: ~{est_seconds/60:.0f} minutes")
    print(f"  (Use --limit N to process a subset first)\n")

    # ── Inference ─────────────────────────────────────────────────────────────
    print(SEP)
    print("  Running inference...")
    print(SEP)

    t_start = time.time()
    results, errors = run_evaluation(entries, endpoint_url, args.timeout)
    elapsed = time.time() - t_start

    # ── Build DataFrames ──────────────────────────────────────────────────────
    if not results:
        sys.exit("\n[ERROR] No results collected — all requests failed.")

    df = pd.DataFrame(results)
    summary_df = build_summary(df)

    # ── Aggregate overall stats ───────────────────────────────────────────────
    n_processed    = len(results)
    n_errors       = len(errors)
    n_correct      = int(df["Correct"].sum())
    accuracy       = n_correct / n_processed * 100
    avg_confidence = float(df["Confidence"].mean())

    overall = {
        "total_images":    n_processed,
        "avg_confidence":  avg_confidence,
        "accuracy":        accuracy,
        "errors":          n_errors,
    }

    # ── Evaluation Summary (user-specified console format) ────────────────────
    inner_sep = "─" * 50
    print(f"\n{SEP}")
    print("  Evaluation Summary")
    print(SEP)

    for _, row in summary_df.iterrows():
        cat  = str(row["Disaster Type"])
        cnt  = int(row["No. of Images"])
        conf = float(row["Average Confidence"])
        acc  = float(row["Accuracy (%)"])
        print(
            f"  {cat:<25} : {cnt:>5} images  "
            f"| Avg Confidence: {conf:5.1f}%  "
            f"| Accuracy: {acc:5.1f}%"
        )

    print(f"\n  {inner_sep}")
    print(f"  Overall Accuracy   : {accuracy:.1f}%")
    print(f"  Total Images       : {n_processed:,}")
    print(f"  Avg Confidence     : {avg_confidence:.1f}%")
    print(f"  Elapsed            : {elapsed/60:.1f} min  ({elapsed/n_processed:.2f}s/image)")
    if n_errors:
        print(f"  Errors             : {n_errors}")

    # ── Save outputs ──────────────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)

    # Drop internal "Correct" boolean and "Category Path" from the exported file
    export_df = df.drop(columns=["Correct", "Category Path"])

    csv_path     = output_dir / "results.csv"
    results_xlsx = output_dir / "results.xlsx"
    summary_xlsx = output_dir / "summary.xlsx"

    print(f"\n  Saving outputs...")
    export_df.to_csv(csv_path, index=False)
    print(f"  {csv_path}")

    write_results_excel(export_df, results_xlsx)
    print(f"  {results_xlsx}")

    write_summary_excel(summary_df, summary_xlsx, overall)
    print(f"  {summary_xlsx}")

    if errors:
        errors_path = output_dir / "errors.csv"
        pd.DataFrame(errors).to_csv(errors_path, index=False)
        print(f"  {errors_path}  ({n_errors} errors)")

    print(f"\n{SEP}")
    print("  Done.")
    print(SEP)
    print()


if __name__ == "__main__":
    main()
